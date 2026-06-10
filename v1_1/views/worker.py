import io
import os
import zipfile

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.db.models import Q
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema
from rest_framework.permissions import IsAuthenticated
from rest_framework.viewsets import ModelViewSet
from v1_1.common_utils.custom_handler import CustomValidationError
from v1_1.common_utils.renderers import XMLRender
from v1_1.common_utils.xml import json_to_xml
from v1_1.models import OrganizationUser, Country
from v1_1.models import Worker, DocumentsWorker, FileDocuments
from v1_1.permissions.owner_or_admin import IsOwnerOrIsAdministratorInOrganization, \
    IsOwnerOrIsAdministratorInOrganizationWorker, IsOwnerOrIsAdministratorForFileDocument
from v1_1.serializers.worker import WorkerSerializer, CreateWorkerSerializer, DocumentsWorkerSerializer, \
    FileDocumentsSerializer, CountrySerializer
from rest_framework import mixins, viewsets, generics
from rest_framework.response import Response
from ..common_utils.renderers import FileRenderer
from rest_framework.renderers import JSONRenderer
from django.db.models import Case, When, Value
from django.db import models
from rest_framework import status


STATUS_LABELS = {
    'vacancy': 'Вакансия',
    'accepted': 'Принят',
    'dismissed': 'Уволен',
}

GENDER_LABELS = {
    'male': 'Мужской',
    'female': 'Женский',
}


@extend_schema(tags=['Worker'])
class ExportWorkersXLSXView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        org_ids = OrganizationUser.objects.filter(user=request.user).values_list('organization_id', flat=True)
        workers = Worker.objects.filter(organization_id__in=org_ids).select_related('organization').order_by(
            'organization__name', 'surname', 'name'
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Сотрудники'

        headers = [
            'Организация', 'Фамилия', 'Имя', 'Отчество', 'Email', 'Телефон',
            'Статус', 'Гражданство', 'Пол', 'Дата рождения', 'Дата трудоустройства',
            'Должность', 'ИНН', 'СНИЛС',
        ]

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row_idx, worker in enumerate(workers, start=2):
            org_name = f"{worker.organization.get_organizational_form_display()} {worker.organization.name}".strip()
            ws.cell(row=row_idx, column=1, value=org_name)
            ws.cell(row=row_idx, column=2, value=worker.surname or '')
            ws.cell(row=row_idx, column=3, value=worker.name or '')
            ws.cell(row=row_idx, column=4, value=worker.patronymic or '')
            ws.cell(row=row_idx, column=5, value=worker.email or '')
            ws.cell(row=row_idx, column=6, value=worker.phone or '')
            ws.cell(row=row_idx, column=7, value=STATUS_LABELS.get(worker.status, worker.status))
            ws.cell(row=row_idx, column=8, value=worker.citizenship or '')
            ws.cell(row=row_idx, column=9, value=GENDER_LABELS.get(worker.gender, worker.gender or ''))
            ws.cell(row=row_idx, column=10, value=str(worker.birthday) if worker.birthday else '')
            ws.cell(row=row_idx, column=11, value=str(worker.date_employment) if worker.date_employment else '')
            ws.cell(row=row_idx, column=12, value=worker.position or '')
            ws.cell(row=row_idx, column=13, value=worker.inn or '')
            ws.cell(row=row_idx, column=14, value=worker.snils or '')

        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="workers.xlsx"'
        return response


@extend_schema(tags=['Worker'])
class CreateWorkerAPIViewSet(mixins.CreateModelMixin, viewsets.GenericViewSet):
    def get_serializer_class(self):
        if self.action in ['create']:
            serializer_class = CreateWorkerSerializer
        else:
            serializer_class = WorkerSerializer

        return serializer_class

    def get_queryset(self):
        # Получение авторизованного пользователя
        user = self.request.user

        # Получение организаций, в которых работает пользователь
        organizations = OrganizationUser.objects.filter(user=user).values_list('organization', flat=True)

        # Фильтрация работников по организации
        queryset = Worker.objects.filter(organization__in=organizations)

        return queryset

    def get_permissions(self):
        # Определение разрешений в зависимости от типа запроса
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            permission_classes = [IsOwnerOrIsAdministratorInOrganization]
        else:
            permission_classes = [IsAuthenticated]

        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Установка поля paid в True перед сохранением
        serializer.validated_data['paid'] = True

        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


@extend_schema(tags=['Worker'])
class UpdateWorkerAPIViewSet(mixins.UpdateModelMixin, mixins.DestroyModelMixin, viewsets.GenericViewSet):
    serializer_class = WorkerSerializer
    permission_class = IsOwnerOrIsAdministratorInOrganizationWorker

    def get_queryset(self):
        # Получение авторизованного пользователя
        user = self.request.user

        # Получение организаций, в которых работает пользователь
        organizations = OrganizationUser.objects.filter(user=user).values_list('organization', flat=True)

        # Фильтрация работников по организации
        queryset = Worker.objects.filter(organization__in=organizations)

        return queryset


@extend_schema(tags=['Worker'])
class ShowWorkersAPIViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    filterset_fields = ['status']
    serializer_class = WorkerSerializer
    permission_class = IsAuthenticated
    renderer_classes = (JSONRenderer, XMLRender)

    def get_queryset(self):
        search = self.request.query_params.get('search', '')
        search_parts = search.split()

        # Получение авторизованного пользователя
        user = self.request.user

        # Получение организации, в которой работает пользователь
        organization_id = self.kwargs.get('organization')

        if not OrganizationUser.objects.filter(user=user, organization=organization_id).exists():
            raise CustomValidationError({'organization': 'Вы не являетесь работником этой организации'})

        # Фильтрация работников по организации
        q_objects = Q(organization=organization_id)

        # Создание Q-объектов для поиска
        for search_part in search_parts:
            q_objects &= (Q(name__icontains=search_part) |
                          Q(surname__icontains=search_part) |
                          Q(patronymic__icontains=search_part))

        # Фильтрация работников по Q-объектам
        workers = Worker.objects.filter(q_objects)
        return workers

    def list_xml(self, request, *args, **kwargs):
        workers = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(workers, many=True)
        xml = json_to_xml(serializer.data)
        return Response(xml)

    def list(self, request, *args, **kwargs):
        if request.query_params.get("format") == "xml":
            return self.list_xml(request, *args, **kwargs)
        return super().list(request, *args, **kwargs)


@extend_schema(tags=['Documents worker'])
class DocumentsWorkerAPIViewSet(ModelViewSet):
    serializer_class = DocumentsWorkerSerializer
    filterset_fields = ['archive']

    def get_queryset(self):
        return DocumentsWorker.objects.filter(Q(worker_id=self.kwargs.get('worker_id')))

    def list(self, request, **kwargs):
        archive = request.query_params.get('archive')

        documents = DocumentsWorker.objects.filter(Q(worker_id=self.kwargs.get('worker_id')))

        # Приводим значение к булевому типу
        if archive is not None:
            archive = archive.lower() == 'true'  # Преобразуем строку в булевое значение
            documents &= DocumentsWorker.objects.filter(Q(archive=archive))

        # Используем Case и When для сортировки по русскому названию типа документа
        documents = documents.annotate(
            type_document_display=Case(
                *[When(type_document=doc_type[0], then=Value(doc_type[1])) for doc_type in
                  DocumentsWorker.TYPES_DOCUMENTS],
                output_field=models.CharField()
            )
        ).order_by('type_document_display')  # Сортировка по русскому названию

        page = self.paginate_queryset(documents)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(documents, many=True)
        return Response(serializer.data)

    def get_permissions(self):
        # Определение разрешений в зависимости от типа запроса
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            permission_classes = [IsOwnerOrIsAdministratorInOrganizationWorker]
        else:
            permission_classes = [IsAuthenticated]

        return [permission() for permission in permission_classes]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        response_data = serializer.create(serializer.validated_data)  # Используем метод create из сериализатора
        return Response(response_data)


@extend_schema(tags=['Documents worker'])
class FileDocumentsAPIViewSet(ModelViewSet):
    serializer_class = FileDocumentsSerializer
    renderer_classes = (JSONRenderer, FileRenderer)

    def get_queryset(self):
        document_id = self.kwargs.get('document_id')
        doc = DocumentsWorker.objects.filter(pk=document_id).first()
        if doc is None:
            return FileDocuments.objects.none()
        user_org_ids = OrganizationUser.objects.filter(
            user=self.request.user
        ).values_list('organization_id', flat=True)
        if doc.worker_id.organization_id not in user_org_ids:
            return FileDocuments.objects.none()
        return FileDocuments.objects.filter(document_id=document_id)

    def get_permissions(self):
        return [IsOwnerOrIsAdministratorForFileDocument()]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        response_data = serializer.create(serializer.validated_data)
        return Response(response_data)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        if request.query_params.get("format") == "zip":
            buffer = io.BytesIO()
            with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                for file_doc in queryset:
                    file_path = file_doc.file_document.path
                    zf.write(file_path, os.path.basename(file_path))
            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="documents.zip"'
            return response

        return super().list(request, *args, **kwargs)


@extend_schema(tags=['Country'])
class ListCountryView(generics.ListAPIView):
    queryset = Country.objects.all()
    permission_classes = ()
    serializer_class = CountrySerializer

